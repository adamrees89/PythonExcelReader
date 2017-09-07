import sqlite3
import openpyxl
import sys
import os

#Create the cell function
def templateCell(s,sn,col,r):
    conn = sqlite3.connect("template.db")
    c = conn.cursor()
    cellRef = col + r
    val = s[cellRef].value
    fname = s[cellRef].font.name
    fsize = s[cellRef].font.size
    fbold = int(s[cellRef].font.bold == 'true')
    fital = int(s[cellRef].font.italic == 'true')
    ccolour = s[cellRef].fill.start_color.index
    data=[cellRef, val, fname, fsize, fbold, fital, ccolour]
    c.execute("INSERT INTO "+sn+" VALUES (?,?,?,?,?,?,?)", data)
    conn.commit()
    conn.close()


'''This function will call the template cell class from its method,
and from init will create the sql table'''

def templateSheet(s):
    conn = sqlite3.connect("template.db")
    c = conn.cursor()
    sn = s.title
    sn = sn.replace(" ", "").replace("+", "").replace("-", "")\
        .replace("/", "").replace("_", "").replace("&", "")\
        .replace("%", "")
    column2 = 'Cell'
    column3 = 'Value'
    column4 = 'Font_Name'
    column5 = 'Font_Size'
    column6 = 'Font_Bold'
    column7 = 'Font_Italic'
    column8 = 'Cell_Colour'
    fieldtype1 = 'INTEGER'
    fieldtype2 = 'TEXT'
    try:
        c.execute('CREATE TABLE {tn}({c2}{ft2}, {c3}{ft2}, {c4}{ft2},'
        ' {c5}{ft2}, {c6}{ft1}, {c7}{ft1}, {c8}{ft1})'
        .format(tn=sn, ft1=fieldtype1, ft2=fieldtype2, c2=column2, 
        c3=column3, c4=column4, c5=column5, c6=column6, c7=column7, 
        c8=column8))
    except sqlite3.Error:
        print('There was a problem with the sql database, is the'
         ' database already open, or does the sheet already exist?')
        sys.exit(5)
    ExtentRow = s.max_row
    ExtentColumn = s.max_column
    rows = list(range(1, ExtentRow))
    column = list(range(1, ExtentColumn))
    for co in column:
        coL = openpyxl.utils.get_column_letter(co)
        for ro in rows:
            templateCell(s, sn, str(coL), str(ro))
    conn.commit()
    conn.close()
